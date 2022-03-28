from openpyxl import Workbook, load_workbook
wb = load_workbook('Practice.xlsx',data_only=True)
ws = wb.active

# # 퀴즈2 점수 변경
# for col in ws.iter_rows(min_row=2) :
#     col[2].value = 10
#
# # H열 Sum을 이용한 총점계산
# for i in range(2,12) :
#     ws['H' + str(i)].value = '=sum(A'+str(i)+':F'+str(i)+')'
#
# wb.save('Practice.xlsx')

# I열 성적정보 추가
for i in range(2,12) :
    total = ws['H' + str(i)]
    grade = ws['I' + str(i)]
    if ws['A'+str(i)].value  < 5 :
        grade.value = 'F'
        continue
    if total.value >= 90 : grade.value = 'A'
    elif total.value >= 80 : grade.value = 'B'
    elif total.value >= 70 : grade.value = 'C'
    else : grade.value = 'D'
wb.save('Practice.xlsx')