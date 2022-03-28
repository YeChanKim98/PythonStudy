from openpyxl import Workbook # 대문자 W임
from random import *
# 기초 준비
wb = Workbook()
ws = wb.active # 활성화 된 시트를 받아옴 - 현재 띄워진 시트를 작업시트로 설정
ws.title = 'Sheet01' # 시트이름 설정

# 시트의 생성방법 : create_sheet(이름,인덱스)의 형태를 가진다
ws = wb.create_sheet('Sheet02',0) # 인자 미작성 시, 기본 명명법, 순차 인덱스로 새로운 시트를 생성한다

# 시트설정
ws.sheet_properties.tabColor = '3399ff' # 해당 시트의 탭 색상을 지정해준다
wb['Sheet01'].sheet_properties.tabColor = '3399ff' # 이름을 통해서 딕셔너리의 형태로 시트에 접근이 가능하다

# 시트 값 입력
ws["A1"] = 'Input Test' # 좌표입력을 통한 값 입력
ws["A2"].value = 'A2' # 벨류태그를 직접 이용
ws.cell(row=3, column=1, value='A3') # cell 함수를 이용

# 시트 값 출력
print(ws["A1"]) # 해당 셀의 정보 출력 - 어느 시트의 어느 좌표이며, 종류는 무엇인지
print(ws["A2"].value) # 지정 셀이 가진 값을 출력. 없으면 None
print(ws.cell(row=3,column=1).value) #  인덱스를 통한 값 출력. / row가 상수, column이 문자

# 시트 복사
# 타겟 파일.copy_worksheet(복사할 원본 시트)
target = wb.copy_worksheet(ws)
target.title='Cop_Sh' # 명명이 없으면 '원본이름 Copy'라는 이름을 가짐
for x in range(1,11) :
    for y in range(1,11) :
        wb['Cop_Sh'].cell(row=x, column=y, value=randint(0,100))

# 저장 및 종료
wb.save('Test_RPA_Xl.xlsx')
print(wb.sheetnames) # 보유 시트 이름 출력
wb.close()