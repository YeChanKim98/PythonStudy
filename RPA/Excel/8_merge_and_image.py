from openpyxl import Workbook
from openpyxl.drawing.image import Image
wb = Workbook()
ws = wb.active

# 셀 병합
ws.merge_cells('B2:D2')
# 병합셀의 값은 항상 좌상단 셀에 입력해야한다
ws['B2'].value = 'I\'m B2'

# 셀 분리
ws.unmerge_cells('B2:D2')
ws['C2'].value = 'I\'m C2'

# 이미지파일 삽입 - 파일, 위치지정이 가능
ws.add_image(Image('img.jpg'),'D1')

wb.save('merge_and_image.xlsx')