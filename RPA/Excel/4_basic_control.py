
# 값을 가져오기 / 행렬 삽입,삭제 / 값 이동

from openpyxl import load_workbook
wb = load_workbook("3_input_cell.xlsx")
ws = wb.active

# iter_rows = 행값을 가져온다.
# iter_cols = 열값을 가져온다.
# 범위값 받아서 출력
for row in ws.iter_rows(min_row=2): # 옵션으로 첫번째 로우는 2번째 줄부터로 지정하여, 제목은 포함하지 않고, 2번째 줄부터 가져온다
    print(row[0].value, end=" ")
    print(row[1].value, end=" ")
    print(row[2].value)

for row in ws.iter_rows(max_row=1) : # 1번째 행까지만 출력. 즉, 속성의 이름을 출력
    # 몇번째 열에 속성 값이 있는지 모르는 경우는, for문을 이용해서 요소를 검색한 후 조건에 합당하면 작동하게 한다
    row[1].value = '랜덤값_1'
    row[2].value = '랜덤값_2'

# .insert_rows(index,len) - len생략가능
# 8번째 열에 len줄 만큼 빈열 추가(엑셀의 Ctrl + '+' 단축키와 동일)
# 같은기능의 .insert_cols(index,len)도 있다
ws.insert_rows(8)

# .delete_rows(index,len) - len생략가능
# insert와 삭제라는 점 이외에 동일.
ws.delete_rows(9)

# move_range(옮길값(좌표값),rows=옮길 행 값,cols=옮길 열 값) : 입력한 이동 값만큼 추가로 좌표가 증가하는 작동
ws.move_range('B1:C11',rows=0,cols=1)
ws['B1'].value = '추가 값'

wb.save("3_input_cell.xlsx")