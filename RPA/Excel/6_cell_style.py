from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import load_workbook
wb = load_workbook('3_input_cell.xlsx')
ws = wb.active

num = ws['A1']
val_1 = ws['B1']
val_2 = ws['C1']

# 셀의 행렬 변경
ws.row_dimensions[1].height = 20
ws.column_dimensions['A'].width = 5
ws.column_dimensions['C'].width = 10

# 셀의 폰트 설정 Font객체에 값을 담아서 지정 셀의 font속성에 적용, 바로 적용하지 않고 변수에 담아서 스타일로 가지고 있는것도 가능
num.font = Font(color='FF0000', size=11, italic=True, bold=True, strike=True) # 색상과 폰트효과 지정
val_1.font = Font(color='AC28B0', size=5 , underline='single')


# 셀 테두리 설정
thin = Border(right=Side(style='thin'),left=Side(style='thin'),bottom=Side(style='thin'),top=Side(style='thin'))
val_1.border = thin
val_2.border = thin

# 조건을 통한 셀 스타일 변경
for row in ws.rows :
    for cell in row :
        cell.alignment = Alignment(horizontal='center', vertical='center') # 셀값의 중앙정렬
        if cell.column == 1 :
            # 셀의 컬럼 순서를 받아와서 1일 경우에는 제목이므로, 생략
            continue
        elif isinstance(cell.value, int) and cell.value > 50 :
            cell.fill = PatternFill(fgColor='66AA66', fill_type='solid')
            cell.font = Font(color='EEEEEE', bold=True)


# 틀 고정 : 지정값을 유동 부분의 좌상단으로 두고 틀을 고정한다.즉 기준이다
ws.freeze_panes='A2'


wb.save('cell_style.xlsx')
