from openpyxl import load_workbook
wb = load_workbook('Test_RPA_Xl.xlsx')
ws = wb.active

# 행렬의 크기를 알 때 출력
for x in range(1,11) :
    for y in range(1,11) :
        print(ws.cell(row=x, column=y).value,end=" ") # 가로로 출력하기 위해 끝문자의 포멧을 \n이 아닌 " "으로 지정함
    print()
print()

# 행렬의 크기를 모를 때 출력 : 값이 입력 된 범위. 즉, 최대 값을 가져와서 +1로 for문에 맞게 범위를 지정해준다
for x in range(1,ws.max_row+1) :
    for y in range(1,ws.max_column+1) :
        print(ws.cell(row=x, column=y).value,end=" ") # 가로로 출력하기 위해 끝문자의 포멧을 \n이 아닌 " "으로 지정함
    print()