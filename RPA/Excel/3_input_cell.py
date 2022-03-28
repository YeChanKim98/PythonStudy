from openpyxl import *
from random import *

wb = Workbook()
ws = wb.active

# value 가 입력된 값을 가져온다면,
# coordinate 는 지정한 셀의 좌표값을 가져온다(A1,V1,B5..)
# 먄악 좌표값이 커지거나 분리해야 할 일이 있으면 변수 = coordinate_from_string(coordinate 값) 을 통해서 튜플형으로 좌표값을 슬라이싱할 수 있다 -> 슬라이싱 후 [0]에는 열 [1]에는 행이 들어간다

# 1줄씩 데이터 입력
ws.append(['번호','값1','값2',]) # 행에 해당 값을 입력(값이 있는 범위에서 맨 뒤 입력)
for i in range(1,11) :
    ws.append([i, randint(0,100), randint(0,100)]) # 앞에 추가한 내용 뒤에 순차적으로 값이 입력된다ㄴ
print()

# B행에 대한 모든 값을 출력
for cell in ws['B'] : # B행에 대한 값이 배열로 저장되어있음
    print(cell.value, end=" ")
print('\n')

# 1열에 대한 모든 값을 출력
for cell in ws[1]:
    print(cell.value,end=" ")
print('\n')

# '행'의 범위를 지정하여 값 출력
for cols in ws['B:C'] : # 슬라이스를 통한 범위지정
    for cell in cols : # 행의 값이 담긴 배열을 for문으로 출력
        print(cell.value)
print()

# '열'의 범위를  지정하여 값 출력
for cell in ws[1:10]:
    for cols in cell :
        print(cols.value,end=" ")
    print()
a={'a':123}

wb.save('3_input_cell.xlsx')