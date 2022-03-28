import datetime
from openpyxl import Workbook, load_workbook
wb = Workbook()
ws = wb.active

ws['A1'] = datetime.datetime.today()
ws['A2'] = "=AVERAGE(B1:B10)"

for i in range(1,11) :
    ws['B'+str(i)] = i

wb.save('formula.xlsx')

# ws.values : 시트의 모든 값을 범위로 지정하여 가져온다. 즉 'ㄱ'자 형태로 입력된 데이터 범위가 있으면. A행의 None값이 끼어서 나온다 - 데이터 범위는 사각형으로 지정하기 때문
# 데이터 받기1 : 아래는 실제 값이 아닌 수식으로 그대로 출력이 된다
# for row in ws.values :
#     for cell in row :
#         print(cell)

# 데이터 받기2 : 파일 로드시, 수식에 대해, 결과 데이터로만 받아오게 인자 부여
# 코드로 생성된 수식의 경우, 여는 순간 작동하여, 데이터가 생기므로, 한번 열었다가 닫아야 None값이 아닌 데이터가 나온다
#  *참고 : 현재 코드는 위에서 실행다가 재생성하므로 보이지 않음. 원래 두파일인데 한파일로 작성해서 그럼
wb = load_workbook('formula.xlsx',data_only=True)
ws = wb.active

for row in ws.values :
     for cell in row :
         print(cell)


